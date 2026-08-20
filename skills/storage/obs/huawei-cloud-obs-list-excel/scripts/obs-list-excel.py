#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
huawei-cloud-obs-list-excel — Query OBS buckets and objects, export to Excel.

Usage:
    python3 obs-list-excel.py --region cn-north-4 [--output /tmp/obs-report.xlsx] [--bucket-type OBJECT] [--max-keys 1000]
"""
import argparse
import json
import os
import sys

# Add scripts dir to path for skill_quality_sdk import
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if _SCRIPT_DIR not in sys.path:
    sys.path.insert(0, _SCRIPT_DIR)

from skill_quality_sdk import quality_report, QualityError


def _validate_output_path(output_path):
    """Validate that the output path is writable before making API calls.
    Raises QualityError if the path cannot be written to.
    """
    parent = os.path.dirname(output_path) or "."
    if not os.access(parent, os.W_OK):
        raise QualityError("U02", f"Output path not writable: {parent}")
    # Check if the file already exists and can be overwritten
    if os.path.exists(output_path):
        if not os.access(output_path, os.W_OK):
            raise QualityError("U02", f"Output file exists but is not writable: {output_path}")


def _detect_credentials():
    """Auto-detect Huawei Cloud credentials from environment variables."""
    ak, sk = None, None
    for key, val in sorted(os.environ.items()):
        upper = key.upper()
        if not (upper.startswith("HUAWEI") or upper.startswith("HW") or upper.startswith("HWC")):
            continue
        if "ACCESS_KEY" in upper or upper.endswith("_AK"):
            ak = val
        if "SECRET_KEY" in upper or upper.endswith("_SK"):
            sk = val
    return ak, sk


def _init_obs_client(region, ak, sk):
    """Initialize OBS SDK client."""
    from huaweicloudsdkcore.auth.credentials import BasicCredentials
    from huaweicloudsdkcore.http.http_config import HttpConfig
    from huaweicloudsdkobs.v1.obs_client import ObsClient

    credentials = BasicCredentials(ak, sk)
    config = HttpConfig.get_default_config()
    config.ignore_ssl_verification = False

    client = ObsClient.new_builder() \
        .with_http_config(config) \
        .with_credentials(credentials) \
        .with_region(region) \
        .build()
    return client


def list_buckets(client, bucket_type=None):
    """List all OBS buckets."""
    from huaweicloudsdkobs.v1.model.list_buckets_request import ListBucketsRequest

    request = ListBucketsRequest()
    if bucket_type:
        request.x_obs_bucket_type = bucket_type

    response = client.list_buckets(request)
    buckets = []
    if response.buckets and response.buckets.bucket:
        for b in response.buckets.bucket:
            buckets.append({
                "name": b.name,
                "creation_date": b.creation_date or "",
                "location": b.location or "",
            })
    return buckets


def list_objects(client, bucket_name, max_keys=1000, prefix=None, marker=None):
    """List objects in a specific OBS bucket."""
    from huaweicloudsdkobs.v1.model.list_objects_request import ListObjectsRequest

    request = ListObjectsRequest()
    request.bucket_name = bucket_name
    request.max_keys = max_keys
    if prefix:
        request.prefix = prefix
    if marker:
        request.marker = marker

    response = client.list_objects(request)
    objects_list = []
    if response.contents:
        for obj in response.contents:
            objects_list.append({
                "key": obj.key,
                "size": obj.size,
                "e_tag": obj.e_tag or "",
                "last_modified": obj.last_modified or "",
                "storage_class": obj.storage_class or "",
            })

    is_truncated = response.is_truncated if hasattr(response, 'is_truncated') else False
    next_marker = response.next_marker if hasattr(response, 'next_marker') else None

    return objects_list, is_truncated, next_marker


def _size_human(size_bytes):
    """Convert bytes to human-readable size."""
    if size_bytes is None:
        return "0 B"
    size = float(size_bytes)
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if abs(size) < 1024.0:
            return f"{size:.2f} {unit}"
        size /= 1024.0
    return f"{size:.2f} PB"


def _build_excel(buckets, bucket_objects, output_path):
    """Build Excel report with bucket and object data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- Sheet 1: Bucket Summary ---
    ws1 = wb.active
    ws1.title = "Buckets Summary"
    ws1.append(["Bucket Name", "Creation Date", "Location", "Object Count", "Total Size"])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for b in buckets:
        name = b["name"]
        objects = bucket_objects.get(name, [])
        obj_count = len(objects)
        total_size = sum((o.get("size", 0) or 0) for o in objects)
        ws1.append([name, b["creation_date"], b["location"], obj_count, _size_human(total_size)])

    for col in ws1.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = max(max_length + 3, 12)

    # --- Sheet 2: All Objects Details ---
    ws2 = wb.create_sheet(title="Objects")
    ws2.append(["Bucket Name", "Object Key", "Size (Bytes)", "Size (Human)", "ETag", "Last Modified", "Storage Class"])

    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for b in buckets:
        name = b["name"]
        objects = bucket_objects.get(name, [])
        for o in objects:
            ws2.append([
                name,
                o["key"],
                o.get("size", 0),
                _size_human(o.get("size", 0)),
                o.get("e_tag", ""),
                o.get("last_modified", ""),
                o.get("storage_class", ""),
            ])

    for col in ws2.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_length + 3, 60)

    wb.save(output_path)
    return output_path


@quality_report(skill_name="huawei-cloud-obs-list-excel", skill_version="1.0.0")
def run(region, output_path, bucket_type, max_keys):
    """Main skill function — list OBS buckets + objects, export to Excel."""
    ak, sk = _detect_credentials()
    if not ak or not sk:
        raise QualityError("C01", "Huawei Cloud credentials not found in environment variables")

    print(f"[OBS List Excel] Initializing client for region: {region}")
    client = _init_obs_client(region, ak, sk)

    print(f"[OBS List Excel] Listing buckets (type={bucket_type or 'all'})...")
    buckets = list_buckets(client, bucket_type or None)
    bucket_objects = {}
    total_objects = 0

    if buckets:
        print(f"[OBS List Excel] Found {len(buckets)} buckets. Listing objects...")
        for b in buckets:
            name = b["name"]
            print(f"  -> Bucket: {name}")
            all_objects = []
            marker = None
            while True:
                objects_list, is_truncated, next_marker = list_objects(
                    client, name, max_keys=max_keys, marker=marker
                )
                all_objects.extend(objects_list)
                if is_truncated and next_marker:
                    marker = next_marker
                else:
                    break
            bucket_objects[name] = all_objects
            total_objects += len(all_objects)
            print(f"       {len(all_objects)} objects listed")
    else:
        print("[OBS List Excel] No buckets found.")

    print(f"[OBS List Excel] Generating Excel report: {output_path}")
    _build_excel(buckets, bucket_objects, output_path)

    print(f"[OBS List Excel] Done. {len(buckets)} buckets, {total_objects} objects -> {output_path}")
    return {"status": "success", "buckets": len(buckets), "objects": total_objects, "output": output_path}


def main():
    parser = argparse.ArgumentParser(description="OBS List to Excel")
    parser.add_argument("--region", required=True, help="Huawei Cloud region (e.g. cn-north-4)")
    parser.add_argument("--output", default="./obs-inventory.xlsx", help="Output Excel file path")
    parser.add_argument("--bucket-type", default="", choices=["OBJECT", "POSIX", ""],
                        help="Filter bucket type: OBJECT (standard), POSIX (parallel filesystem)")
    parser.add_argument("--max-keys", type=int, default=1000, help="Max objects per bucket list call (1-1000)")

    args = parser.parse_args()

    # Validate --max-keys range (OBS limit: 1-1000)
    if args.max_keys < 1 or args.max_keys > 1000:
        parser.error(f"--max-keys must be between 1 and 1000, got {args.max_keys}")

    # Validate output path writability before API calls
    _validate_output_path(args.output)

    result = run(region=args.region, output_path=args.output,
                 bucket_type=args.bucket_type, max_keys=args.max_keys)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()